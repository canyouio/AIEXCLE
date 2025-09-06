import os
import random
import time
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelUtils:
    def __init__(self):
        # 初始化Excel工具类
        pass
    
    def create_excel(self, file_path, sheet_name, data):
        """
        创建一个新的Excel文件
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 写入数据
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, cell_data in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_data)
            
            # 保存文件
            wb.save(file_path)
            print(f"Excel文件已创建: {file_path}")
            return True
        except Exception as e:
            print(f"创建Excel文件失败: {str(e)}")
            return False
    
    def read_excel(self, file_path):
        """
        读取Excel文件内容
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            content = {}
            content['sheet_name'] = ws.title
            content['data'] = []
            
            # 读取所有数据
            for row in ws.iter_rows(values_only=True):
                content['data'].append(list(row))
            
            return content
        except Exception as e:
            print(f"读取Excel文件失败: {str(e)}")
            return {"error": str(e)}
    
    def edit_excel(self, input_path, output_path, operations):
        """
        根据操作列表编辑Excel文件
        """
        try:
            # 加载现有工作簿
            wb = load_workbook(input_path)
            ws = wb.active
            
            # 执行操作
            if isinstance(operations, list):
                for op in operations:
                    if op.get('type') == 'add_row':
                        row_data = op.get('data', [])
                        row_index = op.get('index', ws.max_row + 1)
                        ws.insert_rows(row_index)
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_index, column=col_idx, value=value)
                    
                    elif op.get('type') == 'update_cell':
                        row = op.get('row')
                        col = op.get('col')
                        value = op.get('value')
                        if row and col:
                            ws.cell(row=row, column=col, value=value)
                    
                    elif op.get('type') == 'format_cell':
                        row = op.get('row')
                        col = op.get('col')
                        format_data = op.get('format', {})
                        if row and col:
                            cell = ws.cell(row=row, column=col)
                            # 应用格式
                            if 'font' in format_data:
                                font_data = format_data['font']
                                font = Font(
                                    name=font_data.get('name'),
                                    size=font_data.get('size'),
                                    bold=font_data.get('bold'),
                                    italic=font_data.get('italic'),
                                    color=font_data.get('color')
                                )
                                cell.font = font
                            # 可以添加更多格式设置
            
            # 保存修改后的文件
            wb.save(output_path)
            print(f"Excel文件已编辑并保存: {output_path}")
            return True
        except Exception as e:
            print(f"编辑Excel文件失败: {str(e)}")
            return False
    
    def apply_style(self, file_path, sheet_name, style_data):
        """
        为Excel文件应用样式
        """
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            # 应用样式
            # 这里可以添加更多样式设置逻辑
            
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"应用样式失败: {str(e)}")
            return False
    
    def create_sample_excel(self, file_path, sheet_name="Sample"):
        """
        创建一个示例Excel文件（用于演示）
        """
        # 创建一个包含模拟数据的Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 添加表头
        headers = ["产品名称", "类别", "价格", "库存数量", "销量"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            # 设置表头样式
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加示例数据
        categories = ["电子产品", "办公用品", "家居用品", "服装", "食品"]
        products = {
            "电子产品": ["智能手机", "笔记本电脑", "平板电脑", "耳机", "充电器"],
            "办公用品": ["笔记本", "钢笔", "文件夹", "订书机", "胶带"],
            "家居用品": ["毛巾", "水杯", "闹钟", "纸巾", "垃圾桶"],
            "服装": ["T恤", "牛仔裤", "运动鞋", "外套", "帽子"],
            "食品": ["饼干", "巧克力", "饮料", "水果", "零食"]
        }
        
        for row_idx in range(2, 12):  # 添加10行数据
            category = random.choice(categories)
            product = random.choice(products[category])
            price = round(random.uniform(10, 1000), 2)
            stock = random.randint(0, 1000)
            sales = random.randint(0, stock)
            
            ws.cell(row=row_idx, column=1, value=product)
            ws.cell(row=row_idx, column=2, value=category)
            ws.cell(row=row_idx, column=3, value=price)
            ws.cell(row=row_idx, column=4, value=stock)
            ws.cell(row=row_idx, column=5, value=sales)
        
        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].auto_size = True
        
        # 保存文件
        wb.save(file_path)
        return True

# 添加mock数据生成函数，用于在API不可用时提供模拟响应
def generate_mock_response(description):
    """
    生成模拟的OpenAI响应数据
    """
    # 基于描述生成一些合理的模拟数据
    if "销售" in description or "业绩" in description:
        return {
            "sheet_name": "销售数据",
            "data": [
                ["月份", "销售额", "目标", "达成率"],
                ["1月", 125000, 120000, "104%"],
                ["2月", 132000, 130000, "102%"],
                ["3月", 145000, 140000, "104%"],
                ["4月", 138000, 140000, "99%"],
                ["5月", 152000, 150000, "101%"],
                ["6月", 160000, 155000, "103%"],
                ["总计", 852000, 835000, "102%"]
            ]
        }
    elif "员工" in description or "人事" in description:
        return {
            "sheet_name": "员工信息",
            "data": [
                ["员工ID", "姓名", "部门", "职位", "入职日期", "薪资"],
                ["EMP001", "张三", "技术部", "高级工程师", "2020-01-15", 15000],
                ["EMP002", "李四", "市场部", "市场经理", "2021-03-20", 12000],
                ["EMP003", "王五", "财务部", "会计", "2019-07-10", 10000],
                ["EMP004", "赵六", "人力资源部", "HR专员", "2022-05-05", 8000],
                ["EMP005", "钱七", "销售部", "销售总监", "2018-11-30", 20000]
            ]
        }
    elif "产品" in description or "库存" in description:
        return {
            "sheet_name": "产品库存",
            "data": [
                ["产品ID", "产品名称", "类别", "单价", "库存数量", "供应商"],
                ["P001", "笔记本电脑", "电子产品", 5899, 120, "科技公司A"],
                ["P002", "无线鼠标", "电子产品", 99, 500, "科技公司B"],
                ["P003", "办公椅", "办公用品", 399, 80, "家具公司C"],
                ["P004", "打印纸", "办公用品", 25, 1000, "文具公司D"],
                ["P005", "保温杯", "日用品", 89, 300, "家居公司E"]
            ]
        }
    else:
        # 默认的示例数据
        return {
            "sheet_name": "示例数据",
            "data": [
                ["列1", "列2", "列3"],
                ["数据1", "数据2", "数据3"],
                ["数据4", "数据5", "数据6"],
                ["数据7", "数据8", "数据9"]
            ]
        }


# 添加数据分析报告生成函数，用于在API不可用时提供模拟分析结果
def generate_analysis_report(excel_content):
    """
    生成模拟的Excel数据分析报告
    """
    try:
        # 解析Excel内容
        data = excel_content.get('data', [])
        sheet_name = excel_content.get('sheet_name', '未知工作表')
        
        # 如果数据为空，返回基本报告
        if not data or len(data) < 2:
            return {
                "status": "success",
                "sheet_name": sheet_name,
                "summary": "工作表数据不足，无法进行详细分析。",
                "insights": [],
                "trends": [],
                "anomalies": [],
                "visualization_data": []
            }
        
        # 提取表头和数据
        headers = data[0]
        rows = data[1:]
        
        # 分析报告模板，确保即使没有特定数据也至少返回一些可视化数据
        report = {
            "status": "success",
            "sheet_name": sheet_name,
            "summary": f"该工作表包含{len(headers)}个字段和{len(rows)}条数据记录。",
            "insights": [],
            "trends": [],
            "anomalies": [],
            "visualization_data": []
        }
        
        # 添加一些基础的可视化数据，确保至少有一个图表显示
        report['visualization_data'].append({
            "type": "bar_chart",
            "title": "数据分布概览",
            "data": [
                {"field": "表头数量", "value": len(headers)},
                {"field": "数据行数", "value": len(rows)}
            ],
            "x_axis": "field",
            "y_axis": "value"
        })
        
        # 根据不同的数据类型生成不同的分析
        # 检查是否包含销售相关数据
        sales_related = any(h in ['销售额', '销量', '业绩', '销售', 'revenue', 'sales'] for h in headers)
        time_related = any(h in ['月份', '日期', '时间', 'month', 'date', 'time'] for h in headers)
        
        if sales_related:
            # 生成销售数据相关的分析
            report['insights'].append("数据包含销售相关信息，显示了业务运营情况。")
            
            # 查找销售额列
            sales_col_idx = None
            for i, header in enumerate(headers):
                if header in ['销售额', '销量', '业绩', 'revenue', 'sales']:
                    sales_col_idx = i
                    break
            
            # 如果找到销售额列，生成更多分析
            if sales_col_idx is not None:
                try:
                    # 提取销售额数据（尝试转换为数字）
                    sales_values = []
                    for row in rows:
                        if len(row) > sales_col_idx and row[sales_col_idx]:
                            try:
                                # 处理可能包含百分号或货币符号的值
                                val = str(row[sales_col_idx]).replace('%', '').replace('￥', '').replace('$', '').replace(',', '')
                                sales_values.append(float(val))
                            except:
                                pass
                    
                    if sales_values:
                        total_sales = sum(sales_values)
                        avg_sales = total_sales / len(sales_values)
                        
                        report['insights'].append(f"总销售额/销量: {total_sales:.2f}")
                        report['insights'].append(f"平均销售额/销量: {avg_sales:.2f}")
                        
                        # 生成可视化数据
                        if time_related:
                            # 查找时间列
                            time_col_idx = None
                            for i, header in enumerate(headers):
                                if header in ['月份', '日期', '时间', 'month', 'date', 'time']:
                                    time_col_idx = i
                                    break
                            
                            if time_col_idx is not None and len(rows) > 0:
                                # 生成时间序列可视化数据
                                time_data = []
                                for row in rows:
                                    if len(row) > time_col_idx and len(row) > sales_col_idx:
                                        try:
                                            time_data.append({
                                                "time": str(row[time_col_idx]),
                                                "value": float(str(row[sales_col_idx]).replace('%', '').replace('￥', '').replace('$', '').replace(',', ''))
                                            })
                                        except:
                                            continue
                                
                                if time_data:
                                    report['visualization_data'].append({
                                        "type": "line_chart",
                                        "title": "销售额趋势",
                                        "data": time_data,
                                        "x_axis": "time",
                                        "y_axis": "value"
                                    })
                                    report['trends'].append("数据显示了明显的时间趋势，可以进行季节性分析。")
                except Exception as e:
                    print(f"生成销售分析时出错: {str(e)}")
        
        # 检查是否包含人员相关数据
        if any(h in ['员工', '姓名', '部门', '职位', 'employee', 'name', 'department', 'position'] for h in headers):
            report['insights'].append("数据包含人员相关信息，可以进行人力资源分析。")
            
            # 生成部门分布可视化数据
            dept_col_idx = None
            for i, header in enumerate(headers):
                if header in ['部门', 'department']:
                    dept_col_idx = i
                    break
            
            if dept_col_idx is not None:
                dept_count = {}
                for row in rows:
                    if len(row) > dept_col_idx and row[dept_col_idx]:
                        dept = row[dept_col_idx]
                        dept_count[dept] = dept_count.get(dept, 0) + 1
                
                if dept_count:
                    report['visualization_data'].append({
                        "type": "pie_chart",
                        "title": "部门人员分布",
                        "data": [
                            {"name": dept, "value": count} for dept, count in dept_count.items()
                        ]
                    })
                    report['insights'].append(f"部门数量: {len(dept_count)}")
        
        # 检查是否包含性别相关数据
        gender_col_idx = None
        for i, header in enumerate(headers):
            if header in ['性别', '男/女', 'gender', 'sex']:
                gender_col_idx = i
                break
        
        if gender_col_idx is not None:
            # 统计性别分布
            gender_count = {}
            for row in rows:
                if len(row) > gender_col_idx and row[gender_col_idx]:
                    gender = str(row[gender_col_idx]).strip()
                    # 标准化性别表示
                    if gender in ['男', 'male', 'Male', 'M', 'm']:
                        gender = '男'
                    elif gender in ['女', 'female', 'Female', 'F', 'f']:
                        gender = '女'
                    gender_count[gender] = gender_count.get(gender, 0) + 1
            
            if gender_count:
                report['insights'].append("数据包含性别信息，已生成性别分布分析。")
                # 计算性别比例
                total = sum(gender_count.values())
                for gender, count in gender_count.items():
                    percentage = (count / total) * 100
                    report['insights'].append(f"{gender}性占比: {percentage:.1f}% ({count}人)")
                
                # 生成性别分布可视化数据 - 同时生成饼图和柱状图
                # 饼图用于显示比例
                report['visualization_data'].append({
                    "type": "pie_chart",
                    "title": "性别分布 - 比例图",
                    "data": [
                        {"name": gender, "value": count} for gender, count in gender_count.items()
                    ]
                })
                
                # 柱状图用于直观比较数量
                bar_data = []
                for gender, count in gender_count.items():
                    bar_data.append({
                        "gender": gender,
                        "count": count
                    })
                
                report['visualization_data'].append({
                    "type": "bar_chart",
                    "title": "性别分布 - 数量比较",
                    "data": bar_data,
                    "x_axis": "gender",
                    "y_axis": "count"
                })
        
        # 检查是否包含产品相关数据
        if any(h in ['产品', '库存', '单价', 'product', 'inventory', 'price'] for h in headers):
            report['insights'].append("数据包含产品和库存相关信息，有助于供应链管理分析。")
        
        # 检查是否包含年龄相关数据
        age_col_idx = None
        for i, header in enumerate(headers):
            if header in ['年龄', '岁数', 'age']:
                age_col_idx = i
                break
        
        if age_col_idx is not None:
            # 统计年龄分布
            age_values = []
            for row in rows:
                if len(row) > age_col_idx and row[age_col_idx]:
                    try:
                        age = float(row[age_col_idx])
                        if age > 0 and age < 150:  # 合理的年龄范围
                            age_values.append(age)
                    except:
                        pass
            
            if age_values:
                # 计算年龄统计信息
                min_age = min(age_values)
                max_age = max(age_values)
                avg_age = sum(age_values) / len(age_values)
                
                report['insights'].append("数据包含年龄信息，已生成年龄统计分析。")
                report['insights'].append(f"年龄范围: {min_age:.0f} - {max_age:.0f}岁")
                report['insights'].append(f"平均年龄: {avg_age:.1f}岁")
                
                # 生成年龄分布可视化数据（按年龄段分组）
                # 定义年龄段
                bins = [0, 18, 30, 40, 50, 60, 100]
                labels = ['0-18岁', '19-30岁', '31-40岁', '41-50岁', '51-60岁', '60岁以上']
                age_groups = {label: 0 for label in labels}
                
                for age in age_values:
                    for i in range(len(bins)-1):
                        if bins[i] < age <= bins[i+1]:
                            age_groups[labels[i]] += 1
                            break
                    if age > bins[-1]:
                        age_groups[labels[-1]] += 1
                
                # 过滤掉数量为0的年龄段
                age_data = [{'name': label, 'value': count} for label, count in age_groups.items() if count > 0]
                
                if age_data:
                    report['visualization_data'].append({
                        "type": "pie_chart",
                        "title": "年龄分布",
                        "data": age_data
                    })
        
        # 分析数值型数据的统计信息
        numeric_headers = []
        for i, header in enumerate(headers):
            # 尝试识别可能的数值型列
            if any(keyword in header.lower() for keyword in ['金额', '数量', '价格', '成绩', '分数', 'salary', 'amount', 'price', 'score']):
                numeric_headers.append((i, header))
        
        for col_idx, header in numeric_headers:
            # 尝试提取数值数据
            numeric_values = []
            for row in rows:
                if len(row) > col_idx and row[col_idx]:
                    try:
                        # 处理可能包含百分号或货币符号的值
                        val = str(row[col_idx]).replace('%', '').replace('￥', '').replace('$', '').replace(',', '')
                        numeric_values.append(float(val))
                    except:
                        continue
            
            if len(numeric_values) > 0:
                # 计算统计信息
                min_val = min(numeric_values)
                max_val = max(numeric_values)
                avg_val = sum(numeric_values) / len(numeric_values)
                
                report['insights'].append(f"{header}统计: 最小值={min_val:.2f}, 最大值={max_val:.2f}, 平均值={avg_val:.2f}")
        
        # 添加一些通用的洞察
        if len(rows) > 100:
            report['insights'].append("数据量较大（超过100条记录），建议使用数据透视表进行深入分析。")
        elif len(rows) < 10:
            report['anomalies'].append("数据量较少，可能影响分析结果的可靠性。")
        
        # 添加一些基本统计信息
        report['insights'].append(f"数据字段数量: {len(headers)}")
        report['insights'].append(f"数据记录数量: {len(rows)}")
        
        # 如果没有具体的洞察，添加一些通用信息
        if not report['insights']:
            report['insights'].append("数据结构完整，适合进行进一步分析。")
        
        return report
        
    except Exception as e:
        print(f"生成分析报告时出错: {str(e)}")
        return {
            "status": "error",
            "message": f"生成分析报告时出错: {str(e)}",
            "sheet_name": excel_content.get('sheet_name', '未知工作表'),
            "insights": [],
            "trends": [],
            "anomalies": [],
            "visualization_data": []
        }


# 添加main函数用于测试
if __name__ == "__main__":
    # 测试创建Excel文件
    utils = ExcelUtils()
    test_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "temp", "test.xlsx")
    utils.create_sample_excel(test_file)
    print(f"测试文件已创建: {test_file}")