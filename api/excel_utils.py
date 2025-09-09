import os
import random
import time
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 如果安装了 pyexcelerate，可以用于高速导出，否则使用 openpyxl 作为回退
try:
    from pyexcelerate import Workbook as PyexcelWorkbook
    HAS_PYEXCEL = True
except:
    HAS_PYEXCEL = False

def parse_excel_to_luckysheet(file_path):
    """将 .xlsx 解析为 Luckysheet 可识别的 JSON（每行: {'cell': {col: value}}）"""
    wb = load_workbook(file_path, data_only=False)  # data_only=False 保留公式
    ws = wb.active
    luckysheet_data = []
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row + 1):
        cell_dict = {}
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is not None:
                # 将列索引转换为从0开始的字符串键
                cell_dict[str(c - 1)] = v
        luckysheet_data.append({"cell": cell_dict})
    # TODO: 额外收集合并单元格、公式、样式等元数据（扩展字段）
    return luckysheet_data

def luckysheet_to_xlsx(luckysheet_data, file_name="edited.xlsx"):
    """
    将 Luckysheet JSON 转为 .xlsx 文件并保存到 config.TEMP_DIR
    返回文件名（不含路径或含路径视实现）
    """
    output_file_name = file_name if file_name.endswith(".xlsx") else f"{file_name}.xlsx"
    output_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "temp", output_file_name)
    # 如果使用 pyexcelerate，可以更快写入
    if HAS_PYEXCEL:
        wb = PyexcelWorkbook()
        sheet = wb.new_sheet("Sheet1", data=[[]])  # will set cells manually below
        # pyexcelerate 行列起点为1
        for r_idx, row in enumerate(luckysheet_data, start=1):
            cell_map = row.get("cell", {})
            for c_str, val in cell_map.items():
                c_idx = int(c_str) + 1
                try:
                    sheet.set_cell(r_idx, c_idx, val)
                except Exception:
                    sheet.set_cell(r_idx, c_idx, str(val))
        wb.save(output_path)
    else:
        # fallback to openpyxl
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for r_idx, row in enumerate(luckysheet_data, start=1):
            cell_map = row.get("cell", {})
            for c_str, val in cell_map.items():
                c_idx = int(c_str) + 1
                ws.cell(row=r_idx, column=c_idx, value=val)
        wb.save(output_path)
    return os.path.basename(output_path)

class ExcelUtils:
    def __init__(self):
        # 初始化Excel工具类
        pass
    
    def create_excel(self, file_path, sheet_name, data):
        """
        创建一个新的Excel文件
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 写入数据
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, cell_data in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_data)
            
            # 保存文件
            wb.save(file_path)
            print(f"Excel文件已创建: {file_path}")
            return True
        except Exception as e:
            print(f"创建Excel文件失败: {str(e)}")
            return False
    
    def read_excel(self, file_path):
        """
        读取Excel文件内容
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            content = {}
            content['sheet_name'] = ws.title
            content['data'] = []
            
            # 读取所有数据
            for row in ws.iter_rows(values_only=True):
                content['data'].append(list(row))
            
            return content
        except Exception as e:
            print(f"读取Excel文件失败: {str(e)}")
            return {"error": str(e)}
    
    def edit_excel(self, input_path, output_path, operations):
        """
        根据操作列表编辑Excel文件
        """
        try:
            # 加载现有工作簿
            wb = load_workbook(input_path)
            ws = wb.active
            
            # 执行操作
            if isinstance(operations, list):
                for op in operations:
                    if op.get('type') == 'add_row':
                        row_data = op.get('data', [])
                        row_index = op.get('index', ws.max_row + 1)
                        ws.insert_rows(row_index)
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_index, column=col_idx, value=value)
                    
                    elif op.get('type') == 'update_cell':
                        row = op.get('row')
                        col = op.get('col')
                        value = op.get('value')
                        if row and col:
                            ws.cell(row=row, column=col, value=value)
                    
                    elif op.get('type') == 'format_cell':
                        row = op.get('row')
                        col = op.get('col')
                        format_data = op.get('format', {})
                        if row and col:
                            cell = ws.cell(row=row, column=col)
                            # 应用格式
                            if 'font' in format_data:
                                font_data = format_data['font']
                                font = Font(
                                    name=font_data.get('name'),
                                    size=font_data.get('size'),
                                    bold=font_data.get('bold'),
                                    italic=font_data.get('italic'),
                                    color=font_data.get('color')
                                )
                                cell.font = font
                            # 可以添加更多格式设置
            
            # 保存修改后的文件
            wb.save(output_path)
            print(f"Excel文件已编辑并保存: {output_path}")
            return True
        except Exception as e:
            print(f"编辑Excel文件失败: {str(e)}")
            return False
    
    def apply_style(self, file_path, sheet_name, style_data):
        """
        为Excel文件应用样式
        """
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            # 应用样式
            # 这里可以添加更多样式设置逻辑
            
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"应用样式失败: {str(e)}")
            return False
    
    def create_sample_excel(self, file_path, sheet_name="Sample"):
        """
        创建一个示例Excel文件（用于演示）
        """
        # 创建一个包含模拟数据的Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 添加表头
        headers = ["产品名称", "类别", "价格", "库存数量", "销量"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            # 设置表头样式
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加示例数据
        categories = ["电子产品", "办公用品", "家居用品", "服装", "食品"]
        products = {
            "电子产品": ["智能手机", "笔记本电脑", "平板电脑", "耳机", "充电器"],
            "办公用品": ["笔记本", "钢笔", "文件夹", "订书机", "胶带"],
            "家居用品": ["毛巾", "水杯", "闹钟", "纸巾", "垃圾桶"],
            "服装": ["T恤", "牛仔裤", "运动鞋", "外套", "帽子"],
            "食品": ["饼干", "巧克力", "饮料", "水果", "零食"]
        }
        
        for row_idx in range(2, 12):  # 添加10行数据
            category = random.choice(categories)
            product = random.choice(products[category])
            price = round(random.uniform(10, 1000), 2)
            stock = random.randint(0, 1000)
            sales = random.randint(0, stock)
            
            ws.cell(row=row_idx, column=1, value=product)
            ws.cell(row=row_idx, column=2, value=category)
            ws.cell(row=row_idx, column=3, value=price)
            ws.cell(row=row_idx, column=4, value=stock)
            ws.cell(row=row_idx, column=5, value=sales)
        
        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].auto_size = True
        
        # 保存文件
        wb.save(file_path)
        return True

# 添加mock数据生成函数，用于在API不可用时提供模拟响应
def generate_mock_response(description):
    """
    生成模拟的OpenAI响应数据
    """
    # 基于描述生成一些合理的模拟数据
    if "销售" in description or"业绩" in description:
        return {
            "sheet_name": "销售数据",
            "data": [
                ["月份", "销售额", "目标", "达成率"],
                ["1月", 125000, 120000, "104%"],
                ["2月", 132000, 130000, "102%"],
                ["3月", 145000, 140000, "104%"],
                ["4月", 138000, 140000, "99%"],
                ["5月", 152000, 150000, "101%"],
                ["6月", 160000, 155000, "103%"],
                ["总计", 852000, 835000, "102%"]
            ]
        }
    elif "员工" in description or "人事" in description:
        return {
            "sheet_name": "员工信息",
            "data": [
                ["员工ID", "姓名", "部门", "职位", "入职日期", "薪资"],
                ["EMP001", "张三", "技术部", "高级工程师", "2020-01-15", 15000],
                ["EMP002", "李四", "市场部", "市场经理", "2021-03-20", 12000],
                ["EMP003", "王五", "财务部", "会计", "2019-07-10", 10000],
                ["EMP004", "赵六", "人力资源部", "HR专员", "2022-05-05", 8000],
                ["EMP005", "钱七", "销售部", "销售总监", "2018-11-30", 20000]
            ]
        }
    elif "产品" in description or "库存" in description:
        return {
            "sheet_name": "产品库存",
            "data": [
                ["产品ID", "产品名称", "类别", "单价", "库存数量", "供应商"],
                ["P001", "笔记本电脑", "电子产品", 5899, 120, "科技公司A"],
                ["P002", "无线鼠标", "电子产品", 99, 500, "科技公司B"],
                ["P003", "办公椅", "办公用品", 399, 80, "家具公司C"],
                ["P004", "打印纸", "办公用品", 25, 1000, "文具公司D"],
                ["P005", "保温杯", "日用品", 89, 300, "家居公司E"]
            ]
        }
    else:
        # 默认的示例数据
        return {
            "sheet_name": "示例数据",
            "data": [
                ["列1", "列2", "列3"],
                ["数据1", "数据2", "数据3"],
                ["数据4", "数据5", "数据6"],
                ["数据7", "数据8", "数据9"]
            ]
        }


# 添加数据分析报告生成函数，用于在API不可用时提供模拟分析结果
# 文件路径: backend/excel_utils.py

def generate_analysis_report(excel_content):
    """
    【最终重构版】生成一个更智能、更专注、且总能包含图表的模拟分析报告。
    """
    try:
        data = excel_content.get('data', [])
        sheet_name = excel_content.get('sheet_name', 'Sheet1')
        
        if not data or len(data) < 2:
            return {
                "status": "success", "sheet_name": sheet_name,
                "summary": "数据行数不足，无法进行有意义的分析。",
                "insights": [], "trends": [], "anomalies": [], "visualization_data": []
            }

        headers = data[0]
        rows = data[1:]
        
        report = {
            "status": "success",
            "sheet_name": sheet_name,
            "summary": f"该工作表包含{len(headers)}个字段和{len(rows)}条数据记录。",
            "insights": ["数据结构完整，适合进行初步分析。"],
            "trends": ["数据中未发现明显的时间或类别趋势。"],
            "anomalies": ["未发现明显的数据异常。"],
            "visualization_data": []
        }

        # --- 智能图表生成逻辑 (重构版) ---
        
        # 辅助函数：尝试从行中安全地提取数值
        def get_numeric_value(row, index):
            try:
                # 兼容带货币符号、百分号、逗号的字符串
                val_str = str(row[index]).replace('%', '').replace('￥', '').replace('$', '').replace(',', '')
                return float(val_str)
            except (ValueError, IndexError, TypeError):
                return None

        # 1. 优先寻找时间序列数据来创建折线图
        time_col_idx = -1
        numeric_col_idx_for_time = -1
        
        for i, h in enumerate(headers):
            if h in ['月份', '日期', '时间', '季度', '年份']:
                time_col_idx = i
                break
        
        if time_col_idx != -1:
            for i, h in enumerate(headers):
                if h in ["销售额", "收入", "利润", "数量", "库存", "价格"]:
                    numeric_col_idx_for_time = i
                    break
        
        if time_col_idx != -1 and numeric_col_idx_for_time != -1:
            chart_data = []
            for row in rows:
                time_val = row[time_col_idx]
                numeric_val = get_numeric_value(row, numeric_col_idx_for_time)
                if time_val is not None and numeric_val is not None:
                    chart_data.append({"time": str(time_val), "value": numeric_val})
            
            if chart_data:
                report["visualization_data"].append({
                    "type": "line_chart",
                    "title": f"{headers[numeric_col_idx_for_time]}随'{headers[time_col_idx]}'的变化趋势",
                    "data": chart_data,
                    "x_axis": "time", "y_axis": "value"
                })
                report["trends"] = [f"数据显示了'{headers[numeric_col_idx_for_time]}'随'{headers[time_col_idx]}'变化的明显趋势。"]
                # 成功生成图表后，直接返回报告
                return report

        # 2. 如果没有时间序列，再寻找类别数据来创建饼图或柱状图
        category_col_idx = -1
        numeric_col_idx_for_cat = -1
        
        for i, h in enumerate(headers):
            if h in ["类别", "部门", "产品名称", "地区", "供应商"]:
                category_col_idx = i
                break
        
        if category_col_idx != -1:
            for i, h in enumerate(headers):
                if h in ["库存数量", "销售额", "数量", "人数"]:
                    numeric_col_idx_for_cat = i
                    break
        
        if category_col_idx != -1 and numeric_col_idx_for_cat != -1:
            chart_data_map = {}
            for row in rows:
                category = row[category_col_idx]
                numeric_val = get_numeric_value(row, numeric_col_idx_for_cat)
                if category is not None and numeric_val is not None:
                    chart_data_map[str(category)] = chart_data_map.get(str(category), 0) + numeric_val
            
            # 如果类别少于7个，用饼图更合适
            if 1 < len(chart_data_map) < 7:
                report["visualization_data"].append({
                    "type": "pie_chart",
                    "title": f"按'{headers[category_col_idx]}'划分的'{headers[numeric_col_idx_for_cat]}'分布",
                    "data": [{"name": k, "value": v} for k, v in chart_data_map.items()],
                })
            elif len(chart_data_map) >= 1: # 类别太多或只有一个类别，用柱状图
                chart_data = [{"category": k, "value": v} for k, v in chart_data_map.items()]
                report["visualization_data"].append({
                    "type": "bar_chart",
                    "title": f"各'{headers[category_col_idx]}'的'{headers[numeric_col_idx_for_cat]}'对比",
                    "data": chart_data,
                    "x_axis": "category", "y_axis": "value"
                })
            
            if report["visualization_data"]:
                report["insights"].append(f"数据中'{headers[category_col_idx]}'和'{headers[numeric_col_idx_for_cat]}'存在显著关联，建议关注其分布情况。")
                return report

        # 3. 【备用方案】如果以上所有智能分析都失败了，生成一个最基础的概览图表
        report["visualization_data"].append({
            "type": "bar_chart",
            "title": "数据基础概览",
            "data": [
                {"name": "总字段数 (列)", "value": len(headers)},
                {"name": "总记录数 (行)", "value": len(rows)}
            ],
            "x_axis": "name",
            "y_axis": "value"
        })
        
        return report
        
    except Exception as e:
        # 返回一个结构完整的错误报告
        return {
            "status": "error",
            "message": f"生成模拟报告时发生错误: {str(e)}",
            "sheet_name": excel_content.get('sheet_name', '未知工作表'),
            "summary": "无法生成报告。",
            "insights": [], "trends": [], "anomalies": [], "visualization_data": []
        }

# 添加main函数用于测试
if __name__ == "__main__":
    # 测试创建Excel文件
    utils = ExcelUtils()
    test_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "temp", "test.xlsx")
    utils.create_sample_excel(test_file)
    print(f"测试文件已创建: {test_file}")